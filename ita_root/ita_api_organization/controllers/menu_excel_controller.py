#   Copyright 2022 NEC Corporation
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.

from curses import keyname
from dataclasses import dataclass
import connexion
import openpyxl
from pyrsistent import m

import pytz
import six
import datetime

from common_libs.common import *  # noqa: F403
from common_libs.api import api_filter
from flask import jsonify, session, g
from copy import copy
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import absolute_coordinate
from common_libs.loadtable import *

@api_filter
def get_excel_filter(organization_id, workspace_id, menu, body=None):  # noqa: E501
    """get_excel_filter

    全件のExcelを取得する # noqa: E501

    :param organization_id: OrganizationID
    :type organization_id: str
    :param workspace_id: WorkspaceID
    :type workspace_id: str
    :param menu: メニュー名
    :type menu: str

    :rtype: InlineResponse2004
    """
    try:
        # 言語設定取得
        # 変数定義
        lang = g.LANGUAGE
        
        # make storage directory for excel
        strage_path = os.environ.get('STORAGEPATH')
        excel_dir = strage_path + "/".join([organization_id, workspace_id]) + "/tmp/excel"
        if not os.path.isdir(excel_dir):
            os.makedirs(excel_dir)
            g.applogger.debug("made excel_dir")
        
        # テーブル名
        t_common_menu = 'T_COMN_MENU'
        t_common_column_class = 'T_COMN_COLUMN_CLASS'
        t_common_column_group = 'T_COMN_COLUMN_GROUP'
        t_common_menu_column_link = 'T_COMN_MENU_COLUMN_LINK'
        t_common_menu_table_link = 'T_COMN_MENU_TABLE_LINK'
        
        # DB接続
        objdbca = DBConnectWs(workspace_id)  # noqa: F405
        
        # メニューのカラム情報を取得
        objmenu = load_table.loadTable(objdbca, menu)   # noqa: F405
        result_data = objmenu.rest_filter({})
        
        #### result_code,msg未対応
        result = {
            "result": "result_code", #result_data[0],
            "data": result_data, #result_data[1],
            "message": "msg" #result_data[2]
        }
        
        # 対象メニューを特定するためのIDを取得する
        ret_t_common_menu = objdbca.table_select(t_common_menu, 'WHERE MENU_NAME_REST = %s AND DISUSE_FLAG = %s', [menu, 0])
        if not ret_t_common_menu:
            log_msg_args = [menu]
            api_msg_args = [menu]
            raise AppException("200-00002", log_msg_args, api_msg_args)  # noqa: F405
        menu_id = ret_t_common_menu[0].get('MENU_ID')
        file_name = ret_t_common_menu[0].get('MENU_NAME_' + lang.upper()) + '_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
        file_path = excel_dir + '/' + file_name
        
        # 『メニュー-カラム紐付管理』テーブルから対象のデータを取得
        retList_t_common_menu_column_link = objdbca.table_select(t_common_menu_column_link, 'WHERE MENU_ID = %s AND DISUSE_FLAG = %s', [menu_id, 0])
        if not retList_t_common_menu_column_link:
            log_msg_args = [menu]
            api_msg_args = [menu]
            raise AppException("200-00005", log_msg_args, api_msg_args)  # noqa: F405
        
        # 『カラムクラスマスタ』テーブルからcolumn_typeの一覧を取得
        ret = objdbca.table_select(t_common_column_class, 'WHERE COLUMN_CLASS_NAME LIKE %s AND DISUSE_FLAG = %s', ['%IDColumn%', 0])
        column_class_master = {}
        for recode in ret:
            column_class_master[recode.get('COLUMN_CLASS_ID')] = recode.get('COLUMN_CLASS_NAME')

        # 使用するCOL_GROUP_IDだけを配列に確保しておく
        active_col_group_id = ['']
        for i, dict_menu_column in enumerate(retList_t_common_menu_column_link):
            tmp = dict_menu_column.get('COL_GROUP_ID')
            if tmp is not None and tmp not in active_col_group_id:
                active_col_group_id.append(tmp)
        
        # 『メニュー-テーブル紐付管理』テーブルから対象のデータを取得
        ret = objdbca.table_select(t_common_menu_table_link, 'WHERE MENU_ID = %s AND DISUSE_FLAG = %s', [menu_id, 0])
        if not ret:
            log_msg_args = [menu]
            api_msg_args = [menu]
            raise AppException("200-00003", log_msg_args, api_msg_args)  # noqa: F405
        
        # 登録更新廃止復活フラグを取得する
        menu_table_link_list = []
        menu_table_link_list.append(ret[0].get('ROW_INSERT_FLAG'))
        menu_table_link_list.append(ret[0].get('ROW_UPDATE_FLAG'))
        menu_table_link_list.append(ret[0].get('ROW_DISUSE_FLAG'))
        menu_table_link_list.append(ret[0].get('ROW_REUSE_FLAG'))
        
        # 色の設定
        # 背景色
        fill_wh = PatternFill(fill_type='solid', fgColor='FFFFFF')
        fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
        fill_gr = PatternFill(fill_type='solid', fgColor='D9D9D9')
        
        # 文字色
        font_bl = openpyxl.styles.Font(name='メイリオ', size=8, color='000000')
        font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
        
        # Alignmentの設定
        al_cc = Alignment(horizontal='center', vertical='center')
        al_lt = Alignment(horizontal='left', vertical='top')
        al_ltw = Alignment(horizontal='left', vertical='top', wrapText=True)
        
        # 罫線の設定
        side = Side(border_style="thin", color="000000")
        sideDash = Side(border_style="dashed", color="000000")
        border = Border(left=side, right=side, top=side, bottom=side)
        borderDash = Border(left=side, right=side, top=sideDash, bottom=sideDash)
        
        
        # ワークブックの新規作成と保存
        wb = Workbook()
        
        # マスタ
        msg = g.appmsg.get_api_message('MSG-30012')
        wb.create_sheet(msg)
        # フィルタ条件
        msg = g.appmsg.get_api_message('MSG-30013')
        wb.create_sheet(msg)
        
        # 「マスタ」シート作成
        # 「マスタ」シートの名前の定義をメインのシートの入力規則で使用するため「マスタ」シートから作成する
        name_define_list = make_master_sheet(wb, objdbca, lang, retList_t_common_menu_column_link, column_class_master, menu_table_link_list)
        
        # シートを指定して編集
        ws = wb.active
        
        # ヘッダー部編集
        startRow = 1
        startClm = 4
        
        # 『カラムグループ管理』テーブルからカラムグループ一覧を取得
        ret = objdbca.table_select(t_common_column_group, 'WHERE COL_GROUP_ID IN %s AND DISUSE_FLAG = %s', [active_col_group_id, 0])
        dict_column_group_id = {}
        dict_column_group_id_name = {}
        if ret:
            for recode in ret:
                dict_column_group_id[recode.get('COL_GROUP_ID')] = recode.get('PA_COL_GROUP_ID')
                # IDと名前の辞書も作成する
                dict_column_group_id_name[recode.get('COL_GROUP_ID')] = recode.get('COL_GROUP_NAME_' + lang.upper())
        
        # ヘッダーの階層の深さを調べる
        depth = get_col_group_depth(retList_t_common_menu_column_link, dict_column_group_id)
        
        # エクセルに表示するヘッダー項目を二次元配列に構築する
        excel_header_list = create_excel_headerlist(lang, ws, depth, retList_t_common_menu_column_link, dict_column_group_id_name, dict_column_group_id)
        
        # 1行目（項目名）のヘッダーを作成する
        ws = create_excel_header_firstline(ws, excel_header_list, depth, startRow, startClm, font_wh, al_cc, fill_bl, border)
        
        startRow = startRow + depth
        startDetailRow = startRow + 7
        
        # 2行目以降を作成する
        # 固定部分
        ws = make_template(ws, startRow, font_bl, fill_gr, borderDash, depth)
        
        # エクセルヘッダー部のカラム情報を作成する
        ws = create_column_info(lang, ws, startRow, startClm, retList_t_common_menu_column_link)
        
        # 明細部編集
        # parameterとfileのリストを取得
        mylist = result.get("data")[1]
        
        # ウィンドウ枠の固定
        ws.freeze_panes = ws.cell(row=startRow + 7, column=4).coordinate
        # 入力規則記憶用辞書
        dataVaridationDict = {}
        for param in mylist:
            for k, v in param.items():
                if k == 'parameter':
                    # カラム位置調整用フラグ
                    column_flg = False
                    for i, (key, value) in enumerate(v.items()):
                        if i == 0:
                            ws.cell(row=startRow + 7, column=1).fill = fill_wh
                            ws.cell(row=startRow + 7, column=1, value='')
                            
                            ws.cell(row=startRow + 7, column=2).fill = fill_wh
                            ws.cell(row=startRow + 7, column=2, value='')
                            
                            # 実行処理種別
                            ws.cell(row=startRow + 7, column=3).font = font_bl
                            ws.cell(row=startRow + 7, column=3).alignment = al_cc
                            ws.cell(row=startRow + 7, column=3).border = border
                            ws.cell(row=startRow + 7, column=3, value='-')
                            # 入力規則の設定
                            dv = DataValidation(type='list', formula1='FILTER_ROW_EDIT_BY_FILE')
                            dv.add(ws.cell(row=startRow + 7, column=3))
                            ws.add_data_validation(dv)
                            dataVaridationDict[get_column_letter(3)] = 'FILTER_ROW_EDIT_BY_FILE'
                        
                        if key == 'discard':
                            column_num = 4
                            column_flg = True
                        else:
                            column_num = startClm + i + 1
                            if column_flg:
                                column_num -= 1
                        
                        ws.cell(row=startRow + 7, column=column_num).number_format = openpyxl.styles.numbers.FORMAT_TEXT
                        ws.cell(row=startRow + 7, column=column_num).font = font_bl
                        ws.cell(row=startRow + 7, column=column_num).border = border
                        ws.cell(row=startRow + 7, column=column_num, value=value)
                        if key in name_define_list:
                            dv = DataValidation(type='list', formula1=key)
                            dv.add(ws.cell(row=startRow + 7, column=column_num))
                            ws.add_data_validation(dv)
                            if key not in dataVaridationDict:
                                dataVaridationDict[get_column_letter(column_num)] = key
                    startRow = startRow + 1
        
        # 空行追加処理
        ws = create_blank_line(ws, dataVaridationDict, 10)
        
        
        # 登録が×の列をグレーにする
        # 明細の数を求める
        column_num = ws.max_column + 1
        row_num = ws.max_row - startDetailRow + 1
        for col_i in range(4, column_num):
            for row_j in range(row_num):
                if ws.cell(row=depth + 1, column=col_i).value == '×':
                    # セルの背景色をグレーにする
                    ws.cell(row=startDetailRow + row_j, column=col_i).fill = fill_gr
        
        # フッターを作成する
        ws = create_footer(ws, font_wh, al_lt, al_cc, fill_bl)
        
        # フィルタ条件シートを指定して編集
        # フィルタ条件
        msg = g.appmsg.get_api_message('MSG-30013')
        ws_filter = wb[msg]
        # 出力日時
        msg = g.appmsg.get_api_message('MSG-30014')
        ws_filter['A1'] = msg
        # 廃止
        msg = g.appmsg.get_api_message('MSG-30006')
        ws_filter['D1'] = msg
        ws_filter['A2'] = datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')
        # 廃止含まず
        msg = g.appmsg.get_api_message('MSG-30016')
        ws_filter['D2'] = msg
        
        wb.save(file_path)  # noqa: E303
        
        # 編集してきたエクセルファイルをエンコードする
        wbEncode = file_encode(file_path)  # noqa: F405 F841
        # エンコード後wbは削除する
        os.remove(file_path)
        
        return wbEncode, 200
    
    except Exception as result:
        return result

# 「マスタ」シートを作成する
def make_master_sheet(wb, objdbca, lang, retList_t_common_menu_column_link, column_class_master, menu_table_link_list):  # noqa: E302
    # マスタシートを指定して編集
    msg = g.appmsg.get_api_message('MSG-30012')
    ws_master = wb[msg]
    
    # 開始位置
    startRow_master = 1
    startClm_master = 6
    
    # 実行処理種別
    msg = g.appmsg.get_api_message('MSG-30002')
    ws_master.cell(row=1, column=3, value=msg)
    if menu_table_link_list[0] == '1':
        # 登録
        msg = g.appmsg.get_api_message('MSG-30004')
        ws_master.cell(row=2, column=3, value=msg)
    if menu_table_link_list[1] == '1':
        # 更新
        msg = g.appmsg.get_api_message('MSG-30005')
        ws_master.cell(row=3, column=3, value=msg)
    if menu_table_link_list[2] == '1':
        # 廃止
        msg = g.appmsg.get_api_message('MSG-30006')
        ws_master.cell(row=4, column=3, value=msg)
    if menu_table_link_list[3] == '1':
        # 復活
        msg = g.appmsg.get_api_message('MSG-30007')
        ws_master.cell(row=5, column=3, value=msg)
    
    # 空白行を消して詰める
    for i in reversed(range(1, 5)):
        if ws_master.cell(row=i, column=3).value == None:
            ws_master.delete_rows(i)
    
    # 名前の定義に使用する範囲を求める
    cnt = menu_table_link_list.count('1')
    
    if cnt >= 1:
        msg = g.appmsg.get_api_message('MSG-30012')
        new_range = DefinedName(name='FILTER_ROW_EDIT_BY_FILE', attr_text=msg + '!$C$2:$C$' + str(cnt + 1))
        wb.defined_names.append(new_range)
    
    # 実行処理種別
    msg = g.appmsg.get_api_message('MSG-30002')
    ws_master.cell(row=1, column=4, value=msg)
    # 廃止
    msg = g.appmsg.get_api_message('MSG-30006')
    ws_master.cell(row=1, column=5, value=msg)
    
    name_define_list = []
    for i, dict_menu_column in enumerate(retList_t_common_menu_column_link):
        ws_master.cell(row=startRow_master, column=startClm_master + i, value=dict_menu_column.get('COLUMN_NAME_' + lang.upper()))
        tmp = dict_menu_column.get('COLUMN_CLASS')
        if tmp is not None and tmp in column_class_master:
            name_define = dict_menu_column.get('COLUMN_NAME_REST')
            ref_table_name = dict_menu_column.get('REF_TABLE_NAME')
            ref_col_name = dict_menu_column.get('REF_COL_NAME')
            ref_multi_lang = dict_menu_column.get('REF_MULTI_LANG')
            if ref_table_name is not None:
                if ref_multi_lang == '1':
                    ref_col_name = ref_col_name + '_' + lang.upper()
                ret = objdbca.table_select(ref_table_name, 'WHERE DISUSE_FLAG = %s', [0])
                startCell = None
                for j, recode in enumerate(ret, 1):
                    if j == 1:
                        startCell = ws_master.cell(row=startRow_master + j, column=startClm_master + i).coordinate
                    ws_master.cell(row=startRow_master + j, column=startClm_master + i, value=recode.get(ref_col_name))
                if startCell is not None:
                    msg = g.appmsg.get_api_message('MSG-30012')
                    endCell = ws_master.cell(row=startRow_master + j, column=startClm_master + i).coordinate
                    new_range = DefinedName(name=name_define, attr_text=msg + '!' + absolute_coordinate(startCell) + ':' + absolute_coordinate(endCell))
                    wb.defined_names.append(new_range)
                    name_define_list.append(name_define)
    
    return name_define_list


# エクセル固定部分の作成
def make_template(ws, startRow, font, fill, borderDash, depth):
    # フォント
    font_or = openpyxl.styles.Font(name='メイリオ', size=8, color='E65000', bold=True)
    font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
    
    # 背景色
    fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
    
    # Alignmentの設定
    al_cc = Alignment(horizontal='center', vertical='center')
    al_ltw = Alignment(horizontal='left', vertical='top', wrapText=True)
    al_rt = Alignment(horizontal='right', vertical='top', wrapText=True)
    
    # 罫線
    side = Side(border_style="thin", color="000000")
    sideDash = Side(border_style="dashed", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    border_r = Border(left=side, right=None, top=side, bottom=side)
    border_l = Border(left=None, right=side, top=side, bottom=side)
    borderDash_r = Border(left=side, right=sideDash, top=side, bottom=side)
    borderDash_ltb = Border(left=sideDash, right=side, top=sideDash, bottom=sideDash)
    
    # 固定部分
    # 注意事項
    msg = g.appmsg.get_api_message('MSG-30001')
    ws.cell(row=1, column=1).font = font_wh
    ws.cell(row=1, column=1).alignment = al_cc
    ws.cell(row=1, column=1).fill = fill_bl
    ws.cell(row=1, column=1).border = border_r
    ws.cell(row=1, column=1, value=msg)
    
    ws.cell(row=1, column=2).font = font_wh
    ws.cell(row=1, column=2).alignment = al_cc
    ws.cell(row=1, column=2).fill = fill_bl
    ws.cell(row=1, column=2).border = border_l
    ws.cell(row=1, column=2, value='')
    
    # 実行処理種別
    msg = g.appmsg.get_api_message('MSG-30002')
    ws.cell(row=1, column=3).font = font_wh
    ws.cell(row=1, column=3).alignment = al_cc
    ws.cell(row=1, column=3).fill = fill_bl
    ws.cell(row=1, column=3).border = border
    ws.cell(row=1, column=3, value=msg)
    
    # 固定部分の結合
    alphaOrd = ord('A')
    for i in range(4):
        alphaChr = chr(alphaOrd)
        ws.merge_cells(alphaChr + '1:' + alphaChr + str(depth))
        alphaOrd = alphaOrd + 1
    
    msg = g.appmsg.get_api_message('MSG-30003')
    ws.cell(row=startRow, column=1).font = font
    ws.cell(row=startRow, column=1).fill = fill
    ws.cell(row=startRow, column=1).alignment = al_ltw
    ws.cell(row=startRow, column=1).border = borderDash_r
    ws.cell(row=startRow, column=1, value=msg)
    ws.merge_cells('A' + str(startRow) + ':' + 'A' + str(startRow + 3))
    
    # 登録更新廃止復活
    msg = g.appmsg.get_api_message('MSG-30004')
    ws.cell(row=startRow, column=2).font = font
    ws.cell(row=startRow, column=2).fill = fill
    ws.cell(row=startRow, column=2).alignment = al_cc
    ws.cell(row=startRow, column=2).border = borderDash_ltb
    ws.cell(row=startRow, column=2, value=msg)
    
    msg = g.appmsg.get_api_message('MSG-30005')
    ws.cell(row=startRow + 1, column=2).font = font
    ws.cell(row=startRow + 1, column=2).fill = fill
    ws.cell(row=startRow + 1, column=2).alignment = al_cc
    ws.cell(row=startRow + 1, column=2).border = borderDash_ltb
    ws.cell(row=startRow + 1, column=2, value=msg)
    
    msg = g.appmsg.get_api_message('MSG-30006')
    ws.cell(row=startRow + 2, column=2).font = font
    ws.cell(row=startRow + 2, column=2).fill = fill
    ws.cell(row=startRow + 2, column=2).alignment = al_cc
    ws.cell(row=startRow + 2, column=2).border = borderDash_ltb
    ws.cell(row=startRow + 2, column=2, value=msg)
    
    msg = g.appmsg.get_api_message('MSG-30007')
    ws.cell(row=startRow + 3, column=2).font = font
    ws.cell(row=startRow + 3, column=2).fill = fill
    ws.cell(row=startRow + 3, column=2).alignment = al_cc
    ws.cell(row=startRow + 3, column=2).border = borderDash_ltb
    ws.cell(row=startRow + 3, column=2, value=msg)
    
    # 実行処理種別
    for i in range(4):
        ws.cell(row=startRow + i, column=3).font = font
        ws.cell(row=startRow + i, column=3).fill = fill
        ws.cell(row=startRow + i, column=3).alignment = al_cc
        ws.cell(row=startRow + i, column=3).border = borderDash
        ws.cell(row=startRow + i, column=3, value='')
    
    # 改行
    msg = g.appmsg.get_api_message('MSG-30008')
    ws.cell(row=startRow + 4, column=1).font = font
    ws.cell(row=startRow + 4, column=1).fill = fill
    ws.cell(row=startRow + 4, column=1).alignment = al_ltw
    ws.cell(row=startRow + 4, column=1).border = border
    ws.cell(row=startRow + 4, column=1, value=msg)
    ws.merge_cells('A' + str(startRow + 4) + ':' + 'B' + str(startRow + 4))
    
    ws.cell(row=startRow + 4, column=3).font = font
    ws.cell(row=startRow + 4, column=3).fill = fill
    ws.cell(row=startRow + 4, column=3).alignment = al_ltw
    ws.cell(row=startRow + 4, column=3).border = border
    ws.cell(row=startRow + 4, column=3, value='')
    
    msg = g.appmsg.get_api_message('MSG-30009')
    ws.cell(row=startRow + 5, column=1).font = font
    ws.cell(row=startRow + 5, column=1).fill = fill
    ws.cell(row=startRow + 5, column=1).alignment = al_rt
    ws.cell(row=startRow + 5, column=1).border = border
    ws.cell(row=startRow + 5, column=1, value=msg)
    ws.merge_cells('A' + str(startRow + 5) + ':' + 'B' + str(startRow + 5))
    
    msg = g.appmsg.get_api_message('MSG-30010')
    ws.cell(row=startRow + 5, column=3).font = font_or
    ws.cell(row=startRow + 5, column=3).fill = fill
    ws.cell(row=startRow + 5, column=3).alignment = al_ltw
    ws.cell(row=startRow + 5, column=3).border = border
    ws.cell(row=startRow + 5, column=3, value=msg)
    
    ws.cell(row=startRow + 6, column=1).font = font_wh
    ws.cell(row=startRow + 6, column=1).fill = fill_bl
    ws.cell(row=startRow + 6, column=1).alignment = al_cc
    ws.cell(row=startRow + 6, column=1).border = border_r
    ws.cell(row=startRow + 6, column=1, value='')
    
    ws.cell(row=startRow + 6, column=2).font = font_wh
    ws.cell(row=startRow + 6, column=2).fill = fill_bl
    ws.cell(row=startRow + 6, column=2).alignment = al_cc
    ws.cell(row=startRow + 6, column=2).border = border_l
    ws.cell(row=startRow + 6, column=2, value='')
    
    # 実行処理種別
    msg = g.appmsg.get_api_message('MSG-30002')
    ws.cell(row=startRow + 6, column=3).font = font_wh
    ws.cell(row=startRow + 6, column=3).fill = fill_bl
    ws.cell(row=startRow + 6, column=3).alignment = al_cc
    ws.cell(row=startRow + 6, column=3).border = border
    ws.cell(row=startRow + 6, column=3, value=msg)
    
    return ws


def get_col_group_depth(list_menu_column_link, dict_col_group):  # noqa: E302
    depth = 1
    for dict_menu_column in list_menu_column_link:
        cnt = 1
        col_group_id = dict_menu_column.get('COL_GROUP_ID')
        if col_group_id is not None:
            # 親が無くなるまで繰り返す
            pa_search_flag = True
            while pa_search_flag:
                cnt = cnt + 1
                col_group_id = dict_col_group.get(col_group_id)
                if col_group_id is None:
                    pa_search_flag = False
                    if depth < cnt:
                        depth = cnt
    return depth


def recursive_get_pa_col_group_id(n, id, dict_group_id):
    if n < 1:
        return id
    
    pa_id = dict_group_id.get(id)
    
    return recursive_get_pa_col_group_id(n - 1, pa_id, dict_group_id)  # noqa: E303


# エクセルに表示するヘッダー項目を二次元配列に構築する
def create_excel_headerlist(lang, ws, depth, retList_t_common_menu_column_link, dict_column_group_id_name, dict_column_group_id):
    # 行を階層分追加する
    if depth > 1:
        ws.insert_rows(1, depth - 1)
    
    excel_header_list = [[] for i in range(depth)]
    
    # 表示する項目を二次元配列に構築する
    for i in range(depth):
        for j, dict_menu_column in enumerate(retList_t_common_menu_column_link):
            column_name = dict_menu_column.get('COLUMN_NAME_' + lang.upper())
            msg = g.appmsg.get_api_message('MSG-30015')
            if column_name == msg:
                excel_header_list[depth - 1 - i].insert(0, column_name)
                continue
            if i == 0:
                excel_header_list[depth - 1 - i].append(column_name)
            elif i == 1:
                group_id = dict_menu_column.get('COL_GROUP_ID')
                if group_id is None:
                    excel_header_list[depth - 1 - i].append(column_name)
                else:
                    excel_header_list[depth - 1 - i].append(dict_column_group_id_name.get(group_id))
            else:
                group_id = recursive_get_pa_col_group_id(i - 1, dict_menu_column.get('COL_GROUP_ID'), dict_column_group_id)
                if group_id is None:
                    excel_header_list[depth - 1 - i].append(column_name)
                else:
                    excel_header_list[depth - 1 - i].append(dict_column_group_id_name.get(group_id))
    
    # 親が一番上にくるようにリストを整える
    for i in range(len(excel_header_list[0])):
        for j in range(depth):
            if excel_header_list[j][i] == excel_header_list[depth - 1][i]:
                loopflg = True
                cnt = 1
                # 子が上にある場合、親と値を入れ替える
                while loopflg:
                    if j + cnt >= depth:
                        loopflg = False
                    elif excel_header_list[j][i] != excel_header_list[j + cnt][i]:
                        loopflg = False
                        excel_header_list[j][i] = excel_header_list[j + cnt][i]
                        excel_header_list[j + cnt][i] = excel_header_list[depth - 1][i]
                    cnt += 1
    
    return excel_header_list

# 1行目（項目名）のヘッダーを作成し、結合する
def create_excel_header_firstline(ws, excel_header_list, depth, startRow, startClm, font_wh, al_cc, fill_bl, border):
    # 1行目（項目名）のヘッダーを作成する
    for i, row in enumerate(excel_header_list):
        # 結合開始位置
        startCell = None
        # 前回ループ時の値との比較用
        tmp = None
        for j, data in enumerate(row):
            msg = g.appmsg.get_api_message('MSG-30015')
            if data == msg:
                # 廃止フラグだったら先頭(startClm行)に出力する
                ws.column_dimensions[ws.cell(row=startRow + i, column=startClm).column_letter].width = 2
                ws.cell(row=startRow + i, column=startClm).font = font_wh
                ws.cell(row=startRow + i, column=startClm).alignment = al_cc
                ws.cell(row=startRow + i, column=startClm).fill = fill_bl
                ws.cell(row=startRow + i, column=startClm).border = border
                ws.cell(row=startRow + i, column=startClm, value=data)
                j -= 1
            else:
                ws.cell(row=startRow + i, column=startClm + j).font = font_wh
                ws.cell(row=startRow + i, column=startClm + j).alignment = al_cc
                ws.cell(row=startRow + i, column=startClm + j).fill = fill_bl
                ws.cell(row=startRow + i, column=startClm + j).border = border
                ws.cell(row=startRow + i, column=startClm + j, value=data)
            if tmp == data and startCell is None:
                startCell = ws.cell(row=startRow + i, column=startClm + j - 1).coordinate
            elif tmp != data and startCell is not None:
                # 横軸の結合
                ws.merge_cells(startCell + ':' + ws.cell(row=startRow + i, column=startClm + j - 1).coordinate)
                startCell = None
            tmp = data
    
    # 縦軸の結合
    for i in range(len(excel_header_list[0])):
        # 結合開始位置
        startCell = None
        # 前回ループ時の値との比較用
        tmp = None
        for j in range(depth):
            data = excel_header_list[j][i]
            if tmp == data:
                if startCell is None:
                    startCell = ws.cell(row=startRow + j - 1, column=startClm + i).coordinate
                if j == depth - 1:
                    # 縦軸の結合
                    ws.merge_cells(startCell + ':' + ws.cell(row=startRow + j, column=startClm + i).coordinate)
            elif tmp != data and startCell is not None:
                # 縦軸の結合
                ws.merge_cells(startCell + ':' + ws.cell(row=startRow + j, column=startClm + i).coordinate)
                startCell = None
            tmp = data
    return ws

# エクセルヘッダー部のカラム情報を作成する
def create_column_info(lang, ws, startRow, startClm, retList_t_common_menu_column_link):
    # 文字色
    font_bl = openpyxl.styles.Font(name='メイリオ', size=8, color='000000')
    font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
    
    # 背景色
    fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
    fill_gr = PatternFill(fill_type='solid', fgColor='D9D9D9')
    
    # Alignmentの設定
    al_cc = Alignment(horizontal='center', vertical='center')
    al_ltw = Alignment(horizontal='left', vertical='top', wrapText=True)
    
    # 罫線の設定
    side = Side(border_style="thin", color="000000")
    sideDash = Side(border_style="dashed", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    borderDash = Border(left=side, right=side, top=sideDash, bottom=sideDash)
    
    # カラム位置調整用フラグ
    column_flg = False
    for i, dict_menu_column in enumerate(retList_t_common_menu_column_link):
        column_name = dict_menu_column.get('COLUMN_NAME_' + lang.upper())
        column_num = 0
        
        # 廃止フラグ
        msg = g.appmsg.get_api_message('MSG-30015')
        if column_name == msg:
            column_num = 4
            column_flg = True
        else:
            column_num = startClm + i + 1
            if column_flg:
                column_num -= 1
        
        # 項目名
        ws.cell(row=startRow + 6, column=column_num).font = font_wh
        ws.cell(row=startRow + 6, column=column_num).alignment = al_cc
        ws.cell(row=startRow + 6, column=column_num).fill = fill_bl
        ws.cell(row=startRow + 6, column=column_num).border = border
        ws.cell(row=startRow + 6, column=column_num, value=column_name)
        # 項目名に合わせて幅の調整をする
        column_len = len(str(column_name))
        adjusted_width = (int(column_len) + 2) * 1.2
        ws.column_dimensions[get_column_letter(column_num)].width = adjusted_width
        
        # 最終更新日時はマイクロ秒部分を見えないように幅を設定する
        msg = g.appmsg.get_api_message('MSG-30017')
        if column_name == msg:
            ws.column_dimensions[get_column_letter(column_num)].width = 15.7
        
        # 登録
        # 更新
        column_class = dict_menu_column.get('COLUMN_CLASS')
        auto_input = dict_menu_column.get('AUTO_INPUT')
        input_item = dict_menu_column.get('INPUT_ITEM')
        required_item = dict_menu_column.get('REQUIRED_ITEM')
        tmp = '×'
        if auto_input != '1' and input_item == '1' and required_item == '1':
            tmp = '●'
        elif auto_input != '1' and input_item == '1' and required_item == '0':
            tmp = '○'
        ws.cell(row=startRow, column=column_num).font = font_bl
        ws.cell(row=startRow, column=column_num).fill = fill_gr
        ws.cell(row=startRow, column=column_num).alignment = al_cc
        ws.cell(row=startRow, column=column_num).border = borderDash
        ws.cell(row=startRow, column=column_num, value=tmp)
        
        ws.cell(row=startRow + 1, column=column_num).font = font_bl
        ws.cell(row=startRow + 1, column=column_num).fill = fill_gr
        ws.cell(row=startRow + 1, column=column_num).alignment = al_cc
        ws.cell(row=startRow + 1, column=column_num).border = borderDash
        ws.cell(row=startRow + 1, column=column_num, value=tmp)
        
        tmp = '×'
        # カラムクラスが12（NoteColumn）の場合は任意を設定する
        if column_class == '12':
            tmp = '○'
        # 廃止(×固定)
        ws.cell(row=startRow + 2, column=column_num).font = font_bl
        ws.cell(row=startRow + 2, column=column_num).fill = fill_gr
        ws.cell(row=startRow + 2, column=column_num).alignment = al_cc
        ws.cell(row=startRow + 2, column=column_num).border = borderDash
        ws.cell(row=startRow + 2, column=column_num, value=tmp)
        
        # 復活(×固定)
        ws.cell(row=startRow + 3, column=column_num).font = font_bl
        ws.cell(row=startRow + 3, column=column_num).fill = fill_gr
        ws.cell(row=startRow + 3, column=column_num).alignment = al_cc
        ws.cell(row=startRow + 3, column=column_num).border = borderDash
        ws.cell(row=startRow + 3, column=column_num, value=tmp)
        
        # 改行
        # 改行OK：MultiTextColumn、NoteColumn、SensitiveMultiTextColumnのもの
        # 改行NG：上記以外
        new_line = 'NG'
        if column_class == '2' or column_class == '12' or column_class == '17':
            new_line = 'OK'
        ws.cell(row=startRow + 4, column=column_num).font = font_bl
        ws.cell(row=startRow + 4, column=column_num).fill = fill_gr
        ws.cell(row=startRow + 4, column=column_num).alignment = al_cc
        ws.cell(row=startRow + 4, column=column_num).border = border
        ws.cell(row=startRow + 4, column=column_num, value=new_line)
        
        # その他注意事項
        ws.cell(row=startRow + 5, column=column_num).font = font_bl
        ws.cell(row=startRow + 5, column=column_num).fill = fill_gr
        ws.cell(row=startRow + 5, column=column_num).alignment = al_ltw
        ws.cell(row=startRow + 5, column=column_num).border = border
        ws.cell(row=startRow + 5, column=column_num, value=dict_menu_column.get('DESCRIPTION_' + lang.upper()))
    
    # フィルター設定
    st = ws.cell(row=startRow + 6, column=3).coordinate
    ed = ws.cell(row=startRow + 6, column=ws.max_column).coordinate
    ws.auto_filter.ref = st + ':' + ed
    
    # セルの高さの設定
    ws.row_dimensions[startRow + 4].height = 13.5
    ws.row_dimensions[startRow + 5].height = 85.5
    
    return ws


# フッターを作成する
def create_footer(ws, font_wh, al_lt, al_cc, fill_bl):
    # 罫線
    side = Side(border_style="thin", color="000000")
    border_r = Border(left=side, right=None, top=side, bottom=side)
    border_l = Border(left=None, right=side, top=side, bottom=side)
    border_rl = Border(left=None, right=None, top=side, bottom=side)
    
    # フッター
    footer_row = ws.max_row + 1
    for i in range(ws.max_column):
        if i == 0:
            msg = g.appmsg.get_api_message('MSG-30011')
            ws.cell(row=footer_row, column=1).font = font_wh
            ws.cell(row=footer_row, column=1).alignment = al_lt
            ws.cell(row=footer_row, column=1).fill = fill_bl
            ws.cell(row=footer_row, column=1).border = border_r
            ws.cell(row=footer_row, column=1, value=msg)
        elif i == ws.max_column - 1:
            ws.cell(row=footer_row, column=i + 1).font = font_wh
            ws.cell(row=footer_row, column=i + 1).alignment = al_cc
            ws.cell(row=footer_row, column=i + 1).fill = fill_bl
            ws.cell(row=footer_row, column=i + 1).border = border_l
            ws.cell(row=footer_row, column=i + 1, value='')
        else:
            ws.cell(row=footer_row, column=i + 1).font = font_wh
            ws.cell(row=footer_row, column=i + 1).alignment = al_cc
            ws.cell(row=footer_row, column=i + 1).fill = fill_bl
            ws.cell(row=footer_row, column=i + 1).border = border_rl
            ws.cell(row=footer_row, column=i + 1, value='')
    
    return ws


# 空行を作成する
def create_blank_line(ws, dataVaridationDict, addline):
    # 空行追加処理
    min_col = 1
    min_row = ws.max_row
    max_col = ws.max_column
    max_row = ws.max_row + addline
    shift_col = 0
    shift_row = 1
    for col in range(min_col, max_col + 1):
        for row in range(min_row, max_row + 1):
            # コピー元のコードを作成(column = 1, row = 1 → A1)
            copyFrmCoord = get_column_letter(col) + str(row)
            
            # コピー先のコードを作成
            copyToCol = get_column_letter(col + shift_col)
            copyToCoord = copyToCol + str(row + shift_row)
            
            # コピー先に値をコピー
            if copyToCol == 'C':
                ws[copyToCoord].value = ws[copyFrmCoord].value
            
            if copyToCol in dataVaridationDict:
                dv = DataValidation(type='list', formula1=dataVaridationDict[copyToCol])
                dv.add(copyToCoord)
                ws.add_data_validation(dv)
            
            # 書式がある場合は書式もコピー
            if ws[copyFrmCoord].has_style:
                ws[copyToCoord]._style = copy(ws[copyFrmCoord]._style)

    return ws



@api_filter
def get_excel_format(organization_id, workspace_id, menu):  # noqa: E501
    """get_excel_format

    新規登録用Excelを取得する # noqa: E501

    :param organization_id: OrganizationID
    :type organization_id: str
    :param workspace_id: WorkspaceID
    :type workspace_id: str
    :param menu: メニュー名
    :type menu: str

    :rtype: InlineResponse2004
    """
    try:
        # 言語設定取得
        # 変数定義
        lang = g.LANGUAGE
        
        # make storage directory for excel
        strage_path = os.environ.get('STORAGEPATH')
        excel_dir = strage_path + "/".join([organization_id, workspace_id]) + "/tmp/excel"
        if not os.path.isdir(excel_dir):
            os.makedirs(excel_dir)
            g.applogger.debug("made excel_dir")
        
        # テーブル名
        t_common_menu = 'T_COMN_MENU'
        t_common_column_class = 'T_COMN_COLUMN_CLASS'
        t_common_column_group = 'T_COMN_COLUMN_GROUP'
        t_common_menu_column_link = 'T_COMN_MENU_COLUMN_LINK'
        t_common_menu_table_link = 'T_COMN_MENU_TABLE_LINK'
        
        # DB接続
        objdbca = DBConnectWs(workspace_id)  # noqa: F405
        
        # メニューのカラム情報を取得
        objmenu = load_table.loadTable(objdbca, menu)   # noqa: F405
        result_data = objmenu.rest_filter({})
        
        #### result_code,msg未対応
        result = {
            "result": "result_code", #result_data[0],
            "data": result_data, #result_data[1],
            "message": "msg" #result_data[2]
        }
        
        # 対象メニューを特定するためのIDを取得する
        ret_t_common_menu = objdbca.table_select(t_common_menu, 'WHERE MENU_NAME_REST = %s AND DISUSE_FLAG = %s', [menu, 0])
        if not ret_t_common_menu:
            log_msg_args = [menu]
            api_msg_args = [menu]
            raise AppException("200-00002", log_msg_args, api_msg_args)  # noqa: F405
        menu_id = ret_t_common_menu[0].get('MENU_ID')
        file_name = ret_t_common_menu[0].get('MENU_NAME_' + lang.upper()) + '_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
        file_path = excel_dir + '/' + file_name
        
        # 『メニュー-カラム紐付管理』テーブルから対象のデータを取得
        retList_t_common_menu_column_link = objdbca.table_select(t_common_menu_column_link, 'WHERE MENU_ID = %s AND DISUSE_FLAG = %s', [menu_id, 0])
        if not retList_t_common_menu_column_link:
            log_msg_args = [menu]
            api_msg_args = [menu]
            raise AppException("200-00005", log_msg_args, api_msg_args)  # noqa: F405
        
        # 『カラムクラスマスタ』テーブルからcolumn_typeの一覧を取得
        ret = objdbca.table_select(t_common_column_class, 'WHERE COLUMN_CLASS_NAME LIKE %s AND DISUSE_FLAG = %s', ['%IDColumn%', 0])
        column_class_master = {}
        for recode in ret:
            column_class_master[recode.get('COLUMN_CLASS_ID')] = recode.get('COLUMN_CLASS_NAME')

        # 使用するCOL_GROUP_IDだけを配列に確保しておく
        active_col_group_id = ['']
        for i, dict_menu_column in enumerate(retList_t_common_menu_column_link):
            tmp = dict_menu_column.get('COL_GROUP_ID')
            if tmp is not None and tmp not in active_col_group_id:
                active_col_group_id.append(tmp)
        
        # 『メニュー-テーブル紐付管理』テーブルから対象のデータを取得
        ret = objdbca.table_select(t_common_menu_table_link, 'WHERE MENU_ID = %s AND DISUSE_FLAG = %s', [menu_id, 0])
        if not ret:
            log_msg_args = [menu]
            api_msg_args = [menu]
            raise AppException("200-00003", log_msg_args, api_msg_args)  # noqa: F405
        
        # 登録更新廃止復活フラグを取得する
        menu_table_link_list = []
        menu_table_link_list.append(ret[0].get('ROW_INSERT_FLAG'))
        menu_table_link_list.append(ret[0].get('ROW_UPDATE_FLAG'))
        menu_table_link_list.append(ret[0].get('ROW_DISUSE_FLAG'))
        menu_table_link_list.append(ret[0].get('ROW_REUSE_FLAG'))
        
        # 色の設定
        # 背景色
        fill_wh = PatternFill(fill_type='solid', fgColor='FFFFFF')
        fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
        fill_gr = PatternFill(fill_type='solid', fgColor='D9D9D9')
        
        # 文字色
        font_bl = openpyxl.styles.Font(name='メイリオ', size=8, color='000000')
        font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
        
        # Alignmentの設定
        al_cc = Alignment(horizontal='center', vertical='center')
        al_lt = Alignment(horizontal='left', vertical='top')
        
        # 罫線の設定
        side = Side(border_style="thin", color="000000")
        sideDash = Side(border_style="dashed", color="000000")
        border = Border(left=side, right=side, top=side, bottom=side)
        borderDash = Border(left=side, right=side, top=sideDash, bottom=sideDash)
        
        # ワークブックの新規作成と保存
        wb = Workbook()
        
        # マスタ
        msg = g.appmsg.get_api_message('MSG-30012')
        wb.create_sheet(msg)
        # フィルタ条件
        msg = g.appmsg.get_api_message('MSG-30013')
        wb.create_sheet(msg)
        
        # 「マスタ」シート作成
        # 「マスタ」シートの名前の定義をメインのシートの入力規則で使用するため「マスタ」シートから作成する
        name_define_list = make_master_sheet(wb, objdbca, lang, retList_t_common_menu_column_link, column_class_master, menu_table_link_list)
        
        # シートを指定して編集
        ws = wb.active
        
        # ヘッダー部編集
        startRow = 1
        startClm = 4
        
        # 『カラムグループ管理』テーブルからカラムグループ一覧を取得
        ret = objdbca.table_select(t_common_column_group, 'WHERE COL_GROUP_ID IN %s AND DISUSE_FLAG = %s', [active_col_group_id, 0])
        dict_column_group_id = {}
        dict_column_group_id_name = {}
        if ret:
            for recode in ret:
                dict_column_group_id[recode.get('COL_GROUP_ID')] = recode.get('PA_COL_GROUP_ID')
                # IDと名前の辞書も作成する
                dict_column_group_id_name[recode.get('COL_GROUP_ID')] = recode.get('COL_GROUP_NAME_' + lang.upper())
        
        # ヘッダーの階層の深さを調べる
        depth = get_col_group_depth(retList_t_common_menu_column_link, dict_column_group_id)
        
        # エクセルに表示するヘッダー項目を二次元配列に構築する
        excel_header_list = create_excel_headerlist(lang, ws, depth, retList_t_common_menu_column_link, dict_column_group_id_name, dict_column_group_id)
        
        # 1行目（項目名）のヘッダーを作成する
        ws = create_excel_header_firstline(ws, excel_header_list, depth, startRow, startClm, font_wh, al_cc, fill_bl, border)
        
        startRow = startRow + depth
        startDetailRow = startRow + 7
        
        # 2行目以降を作成する
        # 固定部分
        ws = make_template(ws, startRow, font_bl, fill_gr, borderDash, depth)
        
        # エクセルヘッダー部のカラム情報を作成する
        ws = create_column_info(lang, ws, startRow, startClm, retList_t_common_menu_column_link)
        
        # 明細部編集
        # parameterとfileのリストを取得
        mylist = result.get("data")[1]
        param = mylist[0].get('parameter')
        
        # ウィンドウ枠の固定
        ws.freeze_panes = ws.cell(row=startRow + 7, column=4).coordinate
        
        # 入力規則記憶用辞書
        dataVaridationDict = {}
        # カラム位置調整用フラグ
        column_flg = False
        for i, key in enumerate(param.keys()):
            if i == 0:
                ws.cell(row=startRow + 7, column=1).fill = fill_wh
                ws.cell(row=startRow + 7, column=1, value='')
                
                ws.cell(row=startRow + 7, column=2).fill = fill_wh
                ws.cell(row=startRow + 7, column=2, value='')
                
                # 実行処理種別
                ws.cell(row=startRow + 7, column=3).font = font_bl
                ws.cell(row=startRow + 7, column=3).alignment = al_cc
                ws.cell(row=startRow + 7, column=3).border = border
                ws.cell(row=startRow + 7, column=3, value='-')
                # 入力規則の設定
                dv = DataValidation(type='list', formula1='FILTER_ROW_EDIT_BY_FILE')
                dv.add(ws.cell(row=startRow + 7, column=3))
                ws.add_data_validation(dv)
                dataVaridationDict[get_column_letter(3)] = 'FILTER_ROW_EDIT_BY_FILE'
            
            if key == 'discard':
                column_num = 4
                column_flg = True
            else:
                column_num = startClm + i + 1
                if column_flg:
                    column_num -= 1
            
            ws.cell(row=startRow + 7, column=column_num).number_format = openpyxl.styles.numbers.FORMAT_TEXT
            ws.cell(row=startRow + 7, column=column_num).font = font_bl
            ws.cell(row=startRow + 7, column=column_num).border = border
            ws.cell(row=startRow + 7, column=column_num, value='')
            if key in name_define_list:
                dv = DataValidation(type='list', formula1=key)
                dv.add(ws.cell(row=startRow + 7, column=column_num))
                ws.add_data_validation(dv)
                if key not in dataVaridationDict:
                    dataVaridationDict[get_column_letter(column_num)] = key
        
        # 空行追加処理
        ws = create_blank_line(ws, dataVaridationDict, 9)
        
        # 登録が×の列をグレーにする
        # 明細の数を求める
        column_num = ws.max_column + 1
        row_num = ws.max_row - startDetailRow + 1
        for col_i in range(4, column_num):
            for row_j in range(row_num):
                if ws.cell(row=depth + 1, column=col_i).value == '×':
                    # セルの背景色をグレーにする
                    ws.cell(row=startDetailRow + row_j, column=col_i).fill = fill_gr
        
        # フッターを作成する
        ws = create_footer(ws, font_wh, al_lt, al_cc, fill_bl)
        
        # フィルタ条件シートを指定して編集
        # フィルタ条件
        msg = g.appmsg.get_api_message('MSG-30013')
        ws_filter = wb[msg]
        # 出力日時
        msg = g.appmsg.get_api_message('MSG-30014')
        ws_filter['A1'] = msg
        # 廃止
        msg = g.appmsg.get_api_message('MSG-30006')
        ws_filter['D1'] = msg
        ws_filter['A2'] = datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')
        # 全レコード
        msg = g.appmsg.get_api_message('MSG-30018')
        ws_filter['D2'] = msg
        
        wb.save(file_path)  # noqa: E303
        
        # 編集してきたエクセルファイルをエンコードする
        wbEncode = file_encode(file_path)  # noqa: F405 F841
        # エンコード後wbは削除する
        os.remove(file_path)
        
        return wbEncode, 200
    
    except Exception as result:
        return result


def get_excel_journal(organization_id, workspace_id, menu):  # noqa: E501
    """get_excel_journal

    変更履歴のExcelを取得する # noqa: E501

    :param organization_id: OrganizationID
    :type organization_id: str
    :param workspace_id: WorkspaceID
    :type workspace_id: str
    :param menu: メニュー名
    :type menu: str

    :rtype: InlineResponse2004
    """
    return 'do some magic!'

@api_filter
def post_excel_filter(organization_id, workspace_id, menu, body=None):  # noqa: E501
    """post_excel_filter

    検索条件を指定し、Excelを取得する # noqa: E501

    :param organization_id: OrganizationID
    :type organization_id: str
    :param workspace_id: WorkspaceID
    :type workspace_id: str
    :param menu: メニュー名
    :type menu: str
    :param body: 
    :type body: dict | bytes

    :rtype: InlineResponse2004
    """
    if connexion.request.is_json:
        body = Object.from_dict(connexion.request.get_json())  # noqa: E501
    return 'do some magic!'


def post_excel_maintenance(organization_id, workspace_id, menu, body=None):  # noqa: E501
    """post_excel_maintenance

    Excelでレコードを登録/更新/廃止/復活する # noqa: E501

    :param organization_id: OrganizationID
    :type organization_id: str
    :param workspace_id: WorkspaceID
    :type workspace_id: str
    :param menu: メニュー名
    :type menu: str
    :param body: 
    :type body: dict | bytes

    :rtype: InlineResponse2004
    """
    if connexion.request.is_json:
        body = Object.from_dict(connexion.request.get_json())  # noqa: E501
    return 'do some magic!'
